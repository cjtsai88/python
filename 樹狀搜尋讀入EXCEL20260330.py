# 日期    : 20230810 CJ TSAI
# 目的    : 讀入EXCEL"短期預示排程.xlsx" 中 分頁 "光陽短期預示訂單", "Ecount庫存餘量分析表", "生管BOM" 及 "鐵管BOM" 
#           產生ECOUNT "採購單" EXCEL上傳格式
# 
# 執行方法及資料維護: 
#         1. 生管BOM 及 鐵管BOM   : 生管部門與品技開發 必須隨時更新維護資料 分頁 "生管BOM"(採買一般子件部品) 及 "鐵管BOM" (採買鐵管部品祥興/聯橋...)
#                                  Ecount 品項編碼 一致性
#         2. 光陽短期預示訂單     : 下載最新 光陽短期預示訂單     貼在 短期預示排程.xlsx 分頁 "預示訂單" 
#         3. Ecount庫存餘量分析表 : 下載最新 Ecount庫存餘量分析表 貼在 短期預示排程.xlsx 分頁 "Ecount庫存餘量分析表"
#         4. 程式執行後會 更新分頁資料 "採購單" 及 "採購單_鐵管" 是為 ECOUNT採購單EXCEL上傳格式 上傳ECOUNT系統
#        
# 程式步驟:
#         1. 讀入EXCEL 分頁 光陽 "短期預示訂單", "Ecount庫存餘量分析表", "生管BOM" 及 "鐵管BOM" 為4種 DATAFRAME資料
#         2. 程式執行時 先檢查以下讀入判別相關資料一致性 避免產出錯誤資料
#            2.1 生管BOM 母件號碼與子件號碼 是否重複建資料 
#            2.2 生管BOM 母件號碼與子件號碼 值有相同的行 
#            2.3 光陽短期預示訂單中母件編號是否有"全部"建立於生管BOM 
#            2.4 生管BOM 母件號碼 是否 全部建入 Ecount 品項編碼 (庫存餘量分析表) 的資料中  
#            2.5 鐵管BOM 鐵管尺寸 資料是否重複
#        ???????    2.6 鐵管BOM 鐵管尺寸 是否 全部建入 Ecount 品項編碼(庫存餘量分析表) 的資料中?????
#            以上條件任一不符  程式停止執行 自動跳出
#         3. 以短期預示訂單 母件編號 在 "生管BOM" 搜尋所屬 所有子件編號 依其子件BOM資料計算所需子件物料數量
#            子件所需物料數量　再區分出　一般子件　與　需切鐵管子件　
#         4. 一般子件編號　轉成ECOUNT採購單EXCEL上傳格式　

import pandas as pd
import numpy as np
import time
from datetime import datetime
import datetime
from datetime import date
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.styles import Font
from collections import defaultdict

# log_訊息 設定檔案與 Sheet 名稱
current_path = os.getcwd()
file_path = os.path.join(current_path, 'messages.xlsx')   
sheet_name = 'log_訊息'

# 取得執行時間
execution_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  
# 記錄程式執行開始時間
start_time = time.time()

# 嘗試開啟 Excel，如果不存在則建立
if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path)
    print("✅ 已開啟 Excel 檔案。")
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]  # **刪除舊的 Sheet**
        sheet = wb.create_sheet(sheet_name)
        print(f"⚠️ Sheet '{sheet_name}' 已清空並重新建立。")
    else:
        sheet = wb.create_sheet(sheet_name)
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name
    print("🆕 未找到 Excel 檔案，已建立新檔案與 Sheet。")

# **在第一列寫入執行時間**
sheet.append([f"執行時間: {execution_time}"])  
sheet.append([])  # **加空行，讓標題從第三列開始**

# **定義 Excel 標題**
headers = ["資訊等級"] + [f"資訊{i}" for i in range(0, 16)] + ["附加資訊"]
sheet.append(headers)  # **確保 Sheet 標題正確**
wb.save(file_path)  # 立即存檔，防止標題丟失


def log_訊息(level, data, number0=None, number1=None, number2=None, number3=None, number4=None,
             number5=None, number6=None, number7=None, number8=None, number9=None,
             number10=None, number11=None, number12=None, number13=None, number14=None, number15=None,
             text=None):
    """
    - 如果 `data` 是 DataFrame，則批次寫入
    - 如果 `data` 是字串，則視為單筆寫入
    - 如果 `data` 是 list、Index、tuple，則轉成文字寫入
    - `number0` ~ `number15` 為子件號碼
    - `text` 為附加資訊
    """

    required_columns = [f"資訊{i}" for i in range(0, 20)]

    # ✅ DataFrame ➜ 批次寫入
    if isinstance(data, pd.DataFrame):
        # 自動修正欄位數量
        if len(data.columns) > len(required_columns):
            # 如果超出，多餘的欄位直接丟掉
            data = data.iloc[:, :len(required_columns)]
        elif len(data.columns) < len(required_columns):
            # 如果不足，自動補空白欄位
            missing_columns = len(required_columns) - len(data.columns)
            for i in range(missing_columns):
                data[f"資訊{len(data.columns) + i}"] = ""

        # 重新套用欄位名稱
        data.columns = required_columns

        for _, row in data.iterrows():
            row_data = [level] + row.tolist() + [text if text else ""]
            sheet.append(row_data)

        sheet.append([])
        print("🎉 DataFrame 已成功寫入 Excel！")

    # ✅ list / tuple / Index ➜ 字串化後單筆寫入
    elif isinstance(data, (list, tuple, pd.Index)):
        data_str = ", ".join(map(str, data))
        row_data = [level] + [data_str] + [""] * 15 + [text if text else ""]
        sheet.append(row_data)
        print(f"✅ 單筆資料 '{level}', '{data_str}', '{text}' 已寫入 Excel。")

    # ✅ 單純字串 ➜ 單筆寫入
    elif isinstance(data, str):
        row_data = [level] + [data] + [""] * 15 + [text if text else ""]
        sheet.append(row_data)
        print(f"✅ 單筆資料 '{level}', '{data}', '{text}' 已寫入 Excel。")

    else:
        print(f"⚠️ 無效的輸入型別：{type(data)}，請傳入 DataFrame、字串、list、Index、tuple 等型別。")

    wb.save(file_path)

#比較原始欄位與 讀入資料欄位，找出讀入欄位缺少的原始欄位欄位
def compare_columns(欄位名稱, 原始欄位, 讀入欄位):
    """
    比較原始欄位與 DataFrame 欄位，找出 df 缺少的欄位
    Parameters:
        expected_columns (list): 原始應有的欄位名稱
        df (pd.DataFrame): 已讀入的資料
    """
    #print ("139 "+欄位名稱, 原始欄位)
    missing_columns = [col for col in 原始欄位 if col not in 讀入欄位.columns]
    原始欄位 = pd.DataFrame(原始欄位)
    #print("141 原始欄位", 原始欄位)
    原始欄位 = 原始欄位.T
    讀入欄位 = pd.DataFrame(讀入欄位.columns)
    #print("143 讀入欄位", 讀入欄位)
    讀入欄位 = 讀入欄位.T

    if missing_columns:
        log_訊息("1", "原始"+欄位名稱 + " 欄位")
        log_訊息("1", 原始欄位)
        log_訊息("1", "讀入"+欄位名稱 + " 欄位")
        log_訊息("1", 讀入欄位)
        log_訊息("1 嚴重錯誤 程式中止", "!!!生管BOM_原始欄位 差異 讀入 生管BOM 格式錯誤 請先更正資料 再執行程式執行") 
        log_訊息("1 嚴重錯誤 程式中止", missing_columns)
        print("!!比較", 欄位名稱, "原始欄位 與 讀入欄位")
        print("原始欄位: ", 原始欄位 )
        print("讀入欄位: ", 讀入欄位.columns.tolist())
        print("!!!!!!!原始欄位 差異: ", missing_columns)
        print("!!! 程式不執行 !!!")
        print("讀入 格式錯誤 請先更正資料 再執行程式執行")
        input("##### 按任一鍵 跳出程式 #####")
        print("")
        exit()  # 跳出程式
      
    else:
        print( 欄位名稱 + "原始欄位 與 讀入欄位 相同")
        log_訊息("2 INFO", 欄位名稱 + "原始欄位 與 讀入欄位 相同")
       

# 預示訂單原始欄位
生管BOM_原始欄位_columns = ["原始母件號碼", "母件號碼", "子件號碼", "訂單種類", "流程種類1", "流程種類2", #"流程種類3",
    "使用量", "最低訂購量", "供應商", "收貨倉庫", "管子尺寸", "切管尺寸", "切管比例", "製造單位", 
    "製造機台", "標準工時", "先行天數", "備註1", "備註2", # "test"
]
#生管BOM_原始欄位 = pd.DataFrame(生管BOM_原始欄位_columns)
#P_生管BOM_原始欄位 = 生管BOM_原始欄位.T

#庫存餘量分析表原始欄位
庫存餘量分析表_原始欄位_columns = ['品項編碼', '品項名稱', '供應商名稱', '庫存數量', 
    '交貨時間', '庫存 附加數量', '未進貨(在途量)', '客戶訂單', '生產工單未領出', #"test1"
]
#庫存餘量分析表_原始欄位 = pd.DataFrame(庫存餘量分析表_原始欄位_columns)
#P_庫存餘量分析表_原始欄位 = 庫存餘量分析表_原始欄位.T

# 設定讀入EXCEL檔案名稱及輸出EXCEL檔案名稱
input_file  = 'your_file.xlsx'
output_file = '樹狀展開結果_含需求量.xlsx'

# 讀取生管BOM 主資料（母子件關係）
df = pd.read_excel(input_file, sheet_name="生管BOM", header =1) # header =1 由 生管BOM 第一列 讀取主資料
df = df.dropna(how='all')  # 移除整列全為 NaN 的資料
df.columns = df.columns.str.strip()
#檢查讀入欄位名稱
compare_columns("生管BOM", 生管BOM_原始欄位_columns, df)

dfs = pd.read_excel(input_file, sheet_name=None  , header =1)
df_庫存 = pd.read_excel(input_file, sheet_name="庫存餘量分析表", header =1) # header =1 由 生管BOM 第一列 讀取主資料
df_庫存 = df_庫存.dropna(how='all')  # 移除整列全為 NaN 的資料
df_庫存.columns = df_庫存.columns.str.strip()

#檢查讀入欄位名稱
compare_columns("庫存餘量分析表", 庫存餘量分析表_原始欄位_columns, df_庫存)


#預示訂單原始欄位
預示訂單原始欄位_columns = ["零件號碼", "合計"]
# 讀取預示訂單
forecast_df = pd.read_excel(input_file, sheet_name='預示訂單', header =1)
forecast_df.columns = forecast_df.columns.str.strip()
print(f"208 預示訂單 讀入 訂單筆數：{forecast_df.shape[0]}")
log_訊息("2 INFO", f"讀入預示訂單  {forecast_df.shape[0]} 訂單筆資料")

#檢查讀入欄位名稱
compare_columns("預示訂單原始欄位", 預示訂單原始欄位_columns, forecast_df)


# 檢查 生管BOM 資料是否重複 母件號碼與子件號碼 重複資料
# 使用 duplicated() 函數判斷整個 DataFrame 是否有重複資料
# 找出重複資料的索引
dfs['生管BOM'] = dfs['生管BOM'].dropna(how='all')  # '生管BOM 有空白列資料 移除整列全為 NaN 的資料
duplicate_indices =dfs['生管BOM'][dfs['生管BOM'].duplicated(subset=["原始母件號碼", "母件號碼", "子件號碼"], keep=False)].index
# 列出相應的行
duplicated_data = dfs['生管BOM'].loc[duplicate_indices]
if len(duplicated_data) > 0:
    print("230 生管BOM 資料重複:", duplicate_indices)
    print(duplicated_data)
# 如果有重複資料，則列出重複資料
#if len(duplicated_data) > 0:
#    log_訊息("1", "生管BOM 原始原始母件號碼 原始原始母件號碼 子件號碼 資料重複 非常嚴重 需馬上處理")
#    log_訊息("1", duplicated_data)
#    print("!!!!!!!!!!!!!!!!  錯誤訊息  !!!!!!!!!!!!!!!!")
#    print(duplicated_data)
#    print("以上資料 生管BOM 母件號碼與子件號碼 資料重複")
#    print("(索引 + 3) 為 EXCEL分頁 生管BOM 行號 : ")
#    print("!!! 程式不執行 !!!")
#    print("請先更正 以上資料 再執行程式執行")
#    input("##### 按任一鍵 跳出程式 #####")
#    print("")
#    exit()  # 跳出程式


# 設"品項編碼"為索引 供 df.at 功能取資料用 例如: df_庫存餘量分析表.at[品項編碼, '庫存數量']
df庫存餘量分析表_品項編碼 = dfs["庫存餘量分析表"]['品項編碼']
dfs["庫存餘量分析表"] = dfs["庫存餘量分析表"].set_index('品項編碼')
# 清除NaN 為0 以便運算
df_庫存餘量分析表= dfs['庫存餘量分析表'].fillna(0)


df生管BOM_母件號碼 = dfs['生管BOM']['母件號碼']
df生管BOM = dfs['生管BOM']
#df生管BOM 有相同子件號碼屬不同母件 所以 只留唯一子件號碼資料 搜尋才不發生錯誤引導 以利以子件號碼搜尋 使用量, 供應商 收貨倉庫 ...
df生管BOM_子件號碼_unique = df生管BOM.drop_duplicates(subset=['子件號碼'])
df生管BOM_子件號碼_unique = df生管BOM_子件號碼_unique.set_index('子件號碼')

df生管BOM_columns = dfs['生管BOM'].columns.to_list()
df生管BOM_temp = pd.DataFrame(columns = df生管BOM_columns)
#print("df生管BOM_temp", df生管BOM_temp)
dfs['生管BOM'] = dfs['生管BOM'].set_index('母件號碼')
dfs['生管BOM'] = dfs['生管BOM'].fillna(0)
#print("115 df生管BOM_母件號碼", df生管BOM_母件號碼)

#預示訂單資料
#預示訂單保留第一行需空白 才讀的到'零件號碼'欄位
df預示訂單_零件號碼 = dfs['預示訂單']['零件號碼']
#print("85 df預示訂單_零件號碼 ", df預示訂單_零件號碼)
dfs["預示訂單"] = dfs["預示訂單"].set_index('零件號碼')
dfs['預示訂單'] = dfs['預示訂單'].fillna(0)
#print("123. df預示訂單_零件號碼", df預示訂單_零件號碼)


# 檢查 預示訂單 母件號碼 是否 全部建入 生管BOM 的資料中 
# 如果 預示訂單母件號碼 未建入 生管BOM DATAFRAME"預示訂單未建BOM" 不為空，跳出程式 不執行
預示訂單未建BOM = df預示訂單_零件號碼[~df預示訂單_零件號碼.isin(df生管BOM_母件號碼)]

# 如果 DataFrame 預示訂單未建BOM 不為空，跳出程式 不執行
if not 預示訂單未建BOM.empty:
    log_訊息("1 嚴重錯誤 程式中止", "預示訂單 母件號碼  未建入 生管BOM : ", "非常嚴重 需馬上處理")
    print("預示訂單 母件號碼 未建入 生管BOM :", 預示訂單未建BOM)
    print("")
    print("!!!!!!!!   警示訊息   !!!!!!!")
    print("以上列出 預示訂單 母件號碼 未建入 '生管BOM' 資料 程式不執行")
    print("請先將列出資料建入 '生管BOM' 再執行程式執行")
    input("##### 按任一鍵 跳出程式 #####")
    print("")
    exit()  # 跳出程式


# 檢查 生管BOM 母件號碼 是否 全部建入 ECOUNT 品項編碼 (庫存餘量分析表) 的資料中 
# 如果 生管BOM 母件號碼 未建入 Eocunt "庫存餘量分析表建品項編碼" 不為空，跳出程式 不執行
Ecount未建品項編碼 = df生管BOM_母件號碼[~df生管BOM_母件號碼.isin(df庫存餘量分析表_品項編碼)]
Ecount未建品項編碼 = pd.DataFrame(Ecount未建品項編碼)
#print ("123 Ecount未建品項編碼 ", Ecount未建品項編碼)
# 如果 DataFrame 不為空，跳出程式 不執行
if not Ecount未建品項編碼.empty:
    log_訊息("1 嚴重錯誤 程式中止", "生管BOM 母件號碼 未建Ecount品項編碼 庫存餘量分析表 非常嚴重 需馬上處理")
    log_訊息("1 嚴重錯誤 程式中止", Ecount未建品項編碼)
    print("生管BOM 母件號碼 未建Ecount品項編碼 庫存餘量分析表", Ecount未建品項編碼)
    print("")
    print("!!!!!!!!   警示訊息   !!!!!!!")
    print("以上列出 生管BOM 母件號碼 未建 Ecount 品項編碼 程式不執行")
    print("(索引 + 3) 為 EXCEL分頁 生管BOM 行號 : ")
    print("請先將列出資料建入 'Ecount 品項編碼' 再執行程式執行")
    input("##### 按任一鍵 跳出程式 #####")
    print("")
    exit()  # 跳出程式

df['使用量'] = pd.to_numeric(df['使用量'], errors='coerce').fillna(0)

# 初始化
Finalrow = pd.DataFrame()
展開紀錄_set = set()
已展開子件_set = set()  # 🆕 新增這一行

def expand_components(df, parent_no, qty_path=1, level=0, parent_path=None, source_parent=None, source_qty=1, original_parent=None):
    try:
        global Finalrow
        global 展開紀錄_set
        global 已展開子件_set

        if parent_path is None:
            parent_path = []

        # ✅ 限定資料來源範圍：只處理屬於該 original_parent 的資料
        df = df[df['原始母件號碼'] == original_parent]

        children = df[df['母件號碼'] == parent_no]
        if children.empty:
            return

        for _, row in children.iterrows():
            child_no = row['子件號碼']

            # ✅ 用 (原始母件, 母件, 子件) 作為 key，避免不同母件展開相同子件被誤略
            展開子件_key = (original_parent, parent_no, child_no)
            if 展開子件_key in 已展開子件_set:
                continue
            已展開子件_set.add(展開子件_key)

            usage = row['使用量']
            total_usage = usage * qty_path
            demand_qty = np.ceil(total_usage * source_qty)
    


            path_str = ' > '.join(parent_path + [parent_no])
            展開key = (original_parent, source_parent, child_no, level, path_str)
            if 展開key in 展開紀錄_set:
                continue
            展開紀錄_set.add(展開key)

            row_copy = row.copy()
            row_copy['來源原始母件號碼'] = original_parent
            row_copy['來源母件號碼'] = source_parent
            row_copy['母件路徑'] = path_str
            row_copy['搜尋_子件號碼'] = child_no
            row_copy['來源合計'] = source_qty
            row_copy['層級'] = level
            row_copy['展開後使用量'] = total_usage
            row_copy['光陽訂單需求量'] = demand_qty

            Finalrow = pd.concat([Finalrow, pd.DataFrame([row_copy])], ignore_index=True)

            # ⛔ 避免遞迴死循環
            if child_no == parent_no:
                continue

            # ⏬ 遞迴展開下一層
            expand_components(
                df,
                parent_no=child_no,
                qty_path=total_usage,
                level=level + 1,
                parent_path=parent_path + [parent_no],
                source_parent=source_parent,
                source_qty=source_qty,
                original_parent=original_parent
            )

    except Exception as e:
        log_訊息("1 嚴重錯誤 程式中止", f"expand_components() 展開錯誤：母件號碼 {parent_no}，錯誤：{str(e)}")



# ✅ 讀入預示訂單開始展開每個原始母件
unique_originals = forecast_df['零件號碼'].dropna().unique()

for original_no in unique_originals:
    try:
        source_qtys = forecast_df.loc[forecast_df['零件號碼'] == original_no, '合計'].values
        if len(source_qtys) == 0:
            log_訊息("1 嚴重錯誤 程式中止", f"找不到零件號碼 {original_no} 對應的合計數量，略過")
            continue
        source_qty = source_qtys[0]

        起始母件們 = df[df['原始母件號碼'] == original_no]['母件號碼'].unique()
        for start_parent in 起始母件們:
            expand_components(
                df,
                parent_no=start_parent,
                qty_path=1,
                level=0,
                parent_path=[],
                source_parent=start_parent,
                source_qty=source_qty,
                original_parent=original_no
            )

    except Exception as e:
        log_訊息("1 嚴重錯誤 程式中止", f"展開原始母件號碼 {original_no} 時發生錯誤：{str(e)}")


Finalrow = Finalrow.sort_values(by='原始母件號碼').reset_index(drop=True)


# 找出「管子尺寸」欄位不為空的列 管子尺寸轉成子件號碼以便查詢餘量分析表庫存===
#mask = Finalrow["管子尺寸"].notna() & (Finalrow["管子尺寸"].astype(str).str.strip() != "")
#Finalrow_管子 = Finalrow[mask].copy()
# 建立新列：子件號碼 = 管子尺寸，其餘欄位空白 ===
#new_rows = pd.DataFrame(columns=df.columns)
#new_rows["子件號碼"]     = Finalrow_管子["管子尺寸"]
#new_rows["管子尺寸"]     = Finalrow_管子["管子尺寸"]
#new_rows["光陽訂單需求量"]       = 1
#new_rows["切管比例"]     = Finalrow_管子["切管比例"]
#new_rows["最低訂購量"]    = Finalrow_管子["最低訂購量"]
#new_rows["所需鐵管支數"]  = 0
#new_rows['所需鐵管支數'] = np.ceil(Finalrow_管子["切管比例"] / Finalrow_管子["光陽訂單需求量"]).astype('Int64')

# 在備註寫上來源說明
#new_rows["流程種類1"] = "採購單"
#new_rows["備註"] = "管子尺寸轉成子件號碼以便查詢餘量分析表庫存"
# 合併原始資料與新列 ===
#Finalrow = pd.concat([Finalrow, new_rows], ignore_index=True)






import pandas as pd
import math


def MRP_採購件向下抵扣(df):

    df = df.copy()

    # =========================================================
    # 1️⃣ 基本欄位處理
    # =========================================================
    num_cols = [
        '層級',
        '光陽訂單需求量',
        '展開後使用量',
        '庫存數量',
        '未進貨(在途量)',
        '客戶訂單',
        '生產工單未領出',
        '最低訂購量'
    ]

    for col in num_cols:

        if col not in df.columns:
            df[col] = 0

        df[col] = pd.to_numeric(
            df[col],
            errors='coerce'
        ).fillna(0)

    # =========================================================
    # 2️⃣ 建立欄位
    # =========================================================
    if '建議採購量' not in df.columns:
        df['建議採購量'] = 0

    if '採購量' not in df.columns:
        df['採購量'] = 0

    if '說明' not in df.columns:
        df['說明'] = ''

    df['是否被抵扣'] = 'N'
    df['抵扣前需求量'] = df['光陽訂單需求量']

    # =========================================================
    # 3️⃣ 先建立 每個料號 的 BOM 關係
    #     母件 -> 子件
    # =========================================================
    bom_map = {}

    for _, row in df.iterrows():

        parent = str(row['母件號碼']).strip()
        child = str(row['搜尋_子件號碼']).strip()

        usage = row['展開後使用量']

        level = int(row['層級'])

        if parent == '' or child == '':
            continue

        key = (parent, level)

        if key not in bom_map:
            bom_map[key] = []

        bom_map[key].append({
            'child': child,
            'usage': usage
        })

    # =========================================================
    # 4️⃣ 先計算 第0層 建議採購量
    # =========================================================
    # 原理：
    # 第0層是真正需求來源
    # 後面層級需求都應跟隨父階採購量
    # =========================================================

    level0_mask = (df['層級'] == 0)

    level0_group = (
        df[level0_mask]
        .groupby('品項編碼', as_index=False)
        .agg({
            '光陽訂單需求量': 'sum',
            '庫存數量': 'max',
            '未進貨(在途量)': 'max',
            '客戶訂單': 'max',
            '生產工單未領出': 'max',
            '最低訂購量': 'max'
        })
    )

    purchase_dict = {}

    for _, row in level0_group.iterrows():

        part = row['品項編碼']

        demand = row['光陽訂單需求量']

        stock = row['庫存數量']

        transit = row['未進貨(在途量)']

        customer = row['客戶訂單']

        wip = row['生產工單未領出']

        # =====================================================
        # 🔥 修正核心公式
        #
        # 真正可用量 =
        # 庫存 + 在途 - 客戶訂單 - 工單佔用
        # =====================================================

        available = (
            stock
            + transit
            - customer
            - wip
        )

        suggest_qty = max(
            0,
            demand - available
        )

        suggest_qty = math.ceil(suggest_qty)

        purchase_dict[part] = suggest_qty

    # =========================================================
    # 5️⃣ 逐層往下展開
    # =========================================================
    # 核心概念：
    #
    # 父件採購多少
    # 子件就只需要供應多少
    #
    # 不再使用子件原始需求量
    # =========================================================

    max_level = int(df['層級'].max())

    for lvl in range(max_level + 1):

        current_rows = df[df['層級'] == lvl]

        for _, row in current_rows.iterrows():

            parent_part = str(row['搜尋_子件號碼']).strip()

            if parent_part == '':
                continue

            # 父件建議採購量
            parent_purchase_qty = purchase_dict.get(
                parent_part,
                0
            )

            # 找下一層子件
            child_key = (parent_part, lvl + 1)

            if child_key not in bom_map:
                continue

            child_list = bom_map[child_key]

            for child_info in child_list:

                child_part = child_info['child']

                usage = child_info['usage']

                # =================================================
                # 子件需求 =
                # 父件建議採購量 × 使用量
                # =================================================
                child_need = parent_purchase_qty * usage

                # 找子件庫存資料
                child_rows = df[
                    df['品項編碼'] == child_part
                ]

                if len(child_rows) == 0:
                    continue

                stock = child_rows.iloc[0]['庫存數量']

                transit = child_rows.iloc[0]['未進貨(在途量)']

                customer = child_rows.iloc[0]['客戶訂單']

                wip = child_rows.iloc[0]['生產工單未領出']

                moq = child_rows.iloc[0]['最低訂購量']

                # =================================================
                # 可用量
                # =================================================
                available = (
                    stock
                    + transit
                    - customer
                    - wip
                )

                child_purchase = max(
                    0,
                    child_need - available
                )

                child_purchase = math.ceil(child_purchase)

                # =================================================
                # 儲存結果
                # =================================================
                purchase_dict[child_part] = child_purchase

                # =================================================
                # 回寫 DF
                # =================================================
                mask = (
                    (df['品項編碼'] == child_part)
                )

                df.loc[
                    mask,
                    '建議採購量'
                ] = child_purchase

                # MOQ
                purchase_qty = max(
                    child_purchase,
                    moq
                )

                if child_purchase == 0:
                    purchase_qty = 0

                df.loc[
                    mask,
                    '採購量'
                ] = purchase_qty

                # 說明
                note = (
                    f"由[{parent_part}]"
                    f"建議採購量"
                    f"{parent_purchase_qty}"
                    f"展開"
                )

                old_note = str(
                    df.loc[mask, '說明'].iloc[0]
                )

                if old_note.strip() == '':
                    df.loc[mask, '說明'] = note
                else:
                    df.loc[mask, '說明'] = (
                        old_note + " | " + note
                    )

    # =========================================================
    # 6️⃣ 回寫第0層
    # =========================================================
    for part, qty in purchase_dict.items():

        mask = (df['品項編碼'] == part)

        df.loc[mask, '建議採購量'] = qty

        moq = df.loc[mask, '最低訂購量'].max()

        purchase_qty = max(qty, moq)

        if qty == 0:
            purchase_qty = 0

        df.loc[mask, '採購量'] = purchase_qty

    # =========================================================
    # 7️⃣ 整數化
    # =========================================================
    int_cols = [
        '光陽訂單需求量',
        '建議採購量',
        '採購量'
    ]

    for col in int_cols:

        df[col] = (
            pd.to_numeric(
                df[col],
                errors='coerce'
            )
            .fillna(0)
            .apply(lambda x: int(math.ceil(x)))
        )

    return df



Finalrow_原始展開結果 = Finalrow.copy()


#Finalrow_原始展開結果 = Finalrow.copy()
Finalrow_子件號碼排序 = Finalrow.copy()

# 刪除多餘多個欄位
Finalrow_子件號碼排序 = Finalrow_子件號碼排序.drop(columns=['製造單位','製造機台','標準工時','先行天數'], errors='ignore')  # errors='ignore' 避免欄位不存在時報錯

#以"搜尋_子件號碼"排序
Finalrow_子件號碼排序 = Finalrow_子件號碼排序.sort_values(by='搜尋_子件號碼').reset_index(drop=True)

Finalrow['品項編碼'] = Finalrow['子件號碼']

# 讀取庫存餘量分析表資料
stock_df_raw = pd.read_excel(input_file, sheet_name='庫存餘量分析表', header=None)
header_row = stock_df_raw[stock_df_raw.iloc[:, 0].astype(str).str.contains('品項編碼')].index[0]
stock_df = pd.read_excel(input_file, sheet_name='庫存餘量分析表', skiprows=header_row)
stock_df.columns = stock_df.columns.str.strip()
#for col in ['庫存數量', '未進貨(在途量)', '生產工單未領出']:8/15/2025
for col in ['庫存數量', '未進貨(在途量)', '客戶訂單','生產工單未領出']:    
    stock_df[col] = pd.to_numeric(stock_df[col], errors='coerce').fillna(0)

# 讀取生管BOM 「最低訂購量」資料（以子件號碼對應）
bom_df = pd.read_excel(input_file, sheet_name='生管BOM', header =1)
bom_df.columns = bom_df.columns.str.strip()
Finalrow.columns = Finalrow.columns.str.strip()
#print("bom_df 欄位：", bom_df.columns.tolist())
#print("Finalrow 欄位：", Finalrow.columns.tolist())
#print(bom_df[['子件號碼', '最低訂購量']].head())

if '最低訂購量' in bom_df.columns:
    bom_df['最低訂購量'] = pd.to_numeric(bom_df['最低訂購量'], errors='coerce').fillna(0)
else:
    bom_df['最低訂購量'] = 0


# 合併「庫存餘量分析表」與「最低訂購量」
#Finalrow = Finalrow.merge(stock_df[['品項編碼', '庫存數量', '未進貨(在途量)', '生產工單未領出']], how='left', on='品項編碼') #8/15/2025
Finalrow = Finalrow.merge(stock_df[['品項編碼', '庫存數量', '未進貨(在途量)', '客戶訂單', '生產工單未領出']], how='left', on='品項編碼')
# 只保留光陽訂單需求量不為 0 的資料
#Finalrow = Finalrow[Finalrow['光陽訂單需求量'] > 0]
Finalrow = Finalrow.sort_values(by='品項編碼').reset_index(drop=True)
Finalrow母件路徑 = Finalrow.sort_values(by='母件路徑').reset_index(drop=True)

#print("500", Finalrow.columns)

#取生管BOM 採購單
採購單Finalrow = Finalrow[Finalrow['流程種類1'] == '採購單']

#取生管BOM 管子尺寸 不為零子件號碼
#鐵管Finalrow = 採購單Finalrow[採購單Finalrow['管子尺寸'].notna()].copy()

#生產工單Finalrow = Finalrow[Finalrow['流程種類2'] == '生產工單']

# 計算每個品項編碼的光陽訂單需求量加總
採購單光陽訂單需求量加總 = 採購單Finalrow.groupby('品項編碼', as_index=False)['光陽訂單需求量'].sum().rename(columns={'光陽訂單需求量': '光陽訂單需求量_合計'})

# 對 Finalrow 以「品項編碼」去重複（保留每個品項一筆，保留第一筆資料）
採購單Finalrow_去重複 = 採購單Finalrow.drop_duplicates(subset='品項編碼', keep='first')

# 合併光陽訂單需求量彙總回來（以品項編碼對應）
採購單Finalrow_加總版 = pd.merge(採購單Finalrow_去重複,  採購單光陽訂單需求量加總, on='品項編碼', how='left')

# 將原本的「光陽訂單需求量」欄位改為「光陽訂單需求量_合計」的值
採購單Finalrow_加總版['光陽訂單需求量'] = 採購單Finalrow_加總版['光陽訂單需求量_合計']

# 最後可以選擇是否要刪掉「光陽訂單需求量_合計」欄位
採購單Finalrow_加總版.drop(columns=['光陽訂單需求量_合計'], inplace=True)

# 建議採購量計算
採購單Finalrow_加總版['建議採購量'] = (
    #採購單Finalrow_加總版['光陽訂單需求量'] + 採購單Finalrow_加總版['生產工單未領出'] - 採購單Finalrow_加總版['庫存數量'] - 採購單Finalrow_加總版['未進貨(在途量)'] #8/15/2025
    採購單Finalrow_加總版['光陽訂單需求量'] + 採購單Finalrow_加總版['生產工單未領出'] - 採購單Finalrow_加總版['庫存數量'] + 採購單Finalrow_加總版['客戶訂單'] - 採購單Finalrow_加總版['未進貨(在途量)']
).fillna(0)
採購單Finalrow_加總版['建議採購量'] = 採購單Finalrow_加總版['建議採購量'].apply(lambda x: x if x > 0 else 0)

# 確保合併後 '最低訂購量' 欄位是數字且非空
採購單Finalrow_加總版['最低訂購量'] = pd.to_numeric(採購單Finalrow_加總版['最低訂購量'], errors='coerce').fillna(0)
# 採購量計算（比較最低訂購量）
採購單Finalrow_加總版['採購量'] = 採購單Finalrow_加總版.apply(
    lambda row: max(row['最低訂購量'], row['建議採購量']) if row['建議採購量'] > 0 else 0,
    axis=1
).astype('Int64')

採購單Finalrow_加總版Before = 採購單Finalrow_加總版.copy()
##################################################################
採購單Finalrow_加總版['說明'] = ''
採購單Finalrow_加總版 = MRP_採購件向下抵扣(採購單Finalrow_加總版)

##################################################################



# ✅ 篩選「展開結果」非鐵管 子件號碼  篩選條件「管子尺寸」為空白且「光陽訂單需求量」> 0 的資料
df_non_pipe = 採購單Finalrow_加總版[採購單Finalrow_加總版['管子尺寸'].isna() & (採購單Finalrow_加總版['光陽訂單需求量'] > 0)].copy()


#計算 採購單 交付日期如果供應商工作日是 0 或 空白 自動加10天
today_str = datetime.today().strftime('%Y/%m/%d')
today = datetime.today()
def 計算交付日期(x):
    try:
        # 判斷工作日天數
        if pd.isna(x) or str(x).strip() == '' or int(x) == 0:
            #如果供應商工作日是 0 或 空白 自動加10天
            days = 10
        else:
            days = int(x)        
        # 計算初步交付日
        delivery_date = today + timedelta(days=days)
        # 如果是星期六 (5) 或星期日 (6)，往前調整到星期五 (4)
        if delivery_date.weekday() == 5:  # 星期六
            delivery_date -= timedelta(days=1)
        elif delivery_date.weekday() == 6:  # 星期日
            delivery_date -= timedelta(days=2)
        return delivery_date.strftime('%Y-%m-%d')

    except:
        return '計算交付日期 錯誤'  # 若有錯誤，回傳空字串

# 以供應商工作日 計算 採購單_非鐵管 交付日期
df_non_pipe['交付日期'] = df_non_pipe['供應商工作日'].apply(計算交付日期)


############################################################
#建立「採購單_非鐵管」資料表
df_non_pipe_po = pd.DataFrame({
    "日期": today_str,
    "序號": "1",
    "客戶/供應商編碼": "",
    "客戶/供應商名稱": df_non_pipe['供應商'],
    "供應商簡稱": "",
    "交易類型": "21",
   "承辦人": "",
    "收貨倉庫": df_non_pipe['收貨倉庫'],
    "交付日期": "",
    "品項編碼": df_non_pipe['子件號碼'],
    "品項名稱": "",
    "品項交付日期": "",
    "數量": df_non_pipe['採購量'],
    "單位": "",
    "摘要": "",
    "摘要1": df_non_pipe['備註1'],
    "摘要2": df_non_pipe['備註2']
})
############################################################





print("df_non_pipe_po",df_non_pipe_po )
# 依「客戶/供應商名稱」排序
df_non_pipe_po = df_non_pipe_po.sort_values(by="客戶/供應商名稱").reset_index(drop=True)



####################################################################################
# 提取 採購單_鐵管 資料
####################################################################################
#取採購單Finalrow_加總版['管子尺寸'] 不為空白資料
鐵管Finalrow = 採購單Finalrow_加總版[採購單Finalrow_加總版['管子尺寸'].notna()].copy()
# 只保留光陽訂單需求量不為 0 的資料
鐵管Finalrow = 鐵管Finalrow[鐵管Finalrow['光陽訂單需求量'] > 0]

鐵管Finalrow_original = 鐵管Finalrow.copy()

# 鐵管 建議採購量 計算
鐵管Finalrow['建議採購量'] = (
      鐵管Finalrow['光陽訂單需求量'] 
    + 鐵管Finalrow['生產工單未領出'] 
    - 鐵管Finalrow['庫存數量'] 
    - 鐵管Finalrow['未進貨(在途量)']
).fillna(0)

# =========================================================
# 只要 切管比例 = NaN / 0 / 空白，就一定會炸
# 🔍【加在這裡】切管比例缺失檢查 + log（一定要在算「所需鐵管支數」前）
# =========================================================
mask_切管比例異常 = (
    鐵管Finalrow['切管比例'].isna() |
    (鐵管Finalrow['切管比例'] == 0)
)

切管比例異常資料 = 鐵管Finalrow.loc[
    mask_切管比例異常,
    ['原始母件號碼', '母件號碼', '子件號碼', '管子尺寸', '切管比例']
]

if not 切管比例異常資料.empty:
    print("605 ",切管比例異常資料 )
    log_訊息("1 嚴重錯誤 程式中止", "⚠ 切管比例 計算時是分母 若為 NaN / 0 / 空白 鐵管無法計算所需支數")
    log_訊息("1 嚴重錯誤 程式中止",'原始母件號碼 母件號碼 子件號碼 管子尺寸 切管比例')
    log_訊息("1 嚴重錯誤 程式中止", 切管比例異常資料)

# =========================================================
# 🧮 原本會炸的地方 → 改成安全版本
# =========================================================
鐵管Finalrow['所需鐵管支數'] = pd.NA

mask_可計算 = 鐵管Finalrow['切管比例'] > 0

鐵管Finalrow.loc[mask_可計算, '所需鐵管支數'] = (
    np.ceil(
        鐵管Finalrow.loc[mask_可計算, '建議採購量'] /
        鐵管Finalrow.loc[mask_可計算, '切管比例']
    )
).astype('Int64')


鐵管Finalrow['建議採購量'] = 鐵管Finalrow['建議採購量'].apply(lambda x: x if x > 0 else 0)
鐵管Finalrow['所需鐵管支數'] = np.ceil(鐵管Finalrow['建議採購量'] / 鐵管Finalrow['切管比例']).astype('Int64')

# 將 "管子尺寸" 欄位值寫入到 "品項編碼"
鐵管Finalrow.loc[:, '品項編碼'] = 鐵管Finalrow['管子尺寸']

# 計算每個鐵管資料品項編碼的光陽訂單需求量加總
鐵管光陽訂單需求量加總 = 鐵管Finalrow.groupby('管子尺寸', as_index=False)['所需鐵管支數'].sum().rename(columns={'所需鐵管支數': '所需鐵管支數_合計'})


# 對 Finalrow 以「品項編碼」去重複（保留每個品項一筆，保留第一筆資料）
鐵管Finalrow_去重複 = 鐵管Finalrow.drop_duplicates(subset='管子尺寸', keep='first')
#鐵管Finalrow_去重複 = 鐵管Finalrow.drop_duplicates(subset='子件號碼', keep='first')

# 合併光陽訂單需求量彙總回來（以品項編碼對應）
鐵管Finalrow_加總版 = pd.merge(鐵管Finalrow_去重複, 鐵管光陽訂單需求量加總, on='管子尺寸', how='left')
#鐵管Finalrow_加總版 = pd.merge(鐵管Finalrow_去重複, 鐵管光陽訂單需求量加總, on='子件號碼', how='left')


#以管子尺寸再加總一次 以便相同"管子尺寸"以不同"子件號碼"的光陽訂單需求量加總 2025/09/01
#鐵管Finalrow_加總版 = 鐵管Finalrow_加總版.groupby('管子尺寸', as_index=False)['所需鐵管支數'].sum().rename(columns={'所需鐵管支數': '所需鐵管支數_2nd合計'})
#鐵管Finalrow_加總版_去重複 = 鐵管Finalrow_加總版.drop_duplicates(subset='管子尺寸', keep='first')
#鐵管Finalrow_2nd加總版 = pd.merge(鐵管Finalrow_加總版_去重複, 鐵管Finalrow_加總版, on='管子尺寸',  how='left')





# 將原本的「光陽訂單需求量」欄位改為「光陽訂單需求量_合計」的值
鐵管Finalrow_加總版['所需鐵管支數'] = 鐵管Finalrow_加總版['所需鐵管支數_合計']
鐵管Finalrow_加總版['交付日期'] = 鐵管Finalrow_加總版['供應商工作日'].apply(計算交付日期)

#鐵管Finalrow_2nd加總版['所需鐵管支數'] = 鐵管Finalrow_2nd加總版['所需鐵管支數_2nd合計']
#鐵管Finalrow_2nd加總版['交付日期'] = 鐵管Finalrow_2nd加總版['供應商工作日'].apply(計算交付日期)




#建立　ECOUNT 上傳格式 「採購單_鐵管」資料表
採購單_鐵管 = pd.DataFrame({
    "日期": today_str,
    "序號": "1",
    "客戶/供應商編碼": "",
    "客戶/供應商名稱": 鐵管Finalrow_加總版['供應商'],
    "供應商簡稱": "",
    "交易類型": "21",
    "承辦人": "",
    "收貨倉庫": 鐵管Finalrow_加總版['收貨倉庫'],
    "交付日期": 鐵管Finalrow_加總版['交付日期'],
    "品項編碼": 鐵管Finalrow_加總版['管子尺寸'],
    "品項名稱": "",
    #"交付日期": 鐵管Finalrow_加總版['交付日期'],
    "交付日期": "",
    "數量": 鐵管Finalrow_加總版['所需鐵管支數'],
    "單位": "",
    "摘要": "",
    "摘要1": 鐵管Finalrow_加總版['備註1'],
    "摘要2": 鐵管Finalrow_加總版['備註2'],
    "最低訂購量": 鐵管Finalrow_加總版['最低訂購量']# 計算參考用
})
# 依「客戶/供應商名稱」排序
採購單_鐵管 = 採購單_鐵管.sort_values(by="客戶/供應商名稱").reset_index(drop=True)

採購單_鐵管_merge_stock_df = 採購單_鐵管.copy()

# 合併庫存餘量分析表資料
採購單_鐵管_merge_stock_df = 採購單_鐵管_merge_stock_df.merge(stock_df[['品項編碼', '庫存數量', '未進貨(在途量)', '生產工單未領出']], how='left', on='品項編碼')
#增加計算欄位鐵管建議採購量
採購單_鐵管_merge_stock_df["鐵管建議採購量"]=0

# 建議採購量計算
採購單_鐵管_merge_stock_df["鐵管建議採購量"] = (
    採購單_鐵管_merge_stock_df['數量'] + 採購單_鐵管_merge_stock_df['生產工單未領出'] - 採購單_鐵管_merge_stock_df['庫存數量'] - 採購單_鐵管_merge_stock_df['未進貨(在途量)']
).fillna(0)
採購單_鐵管_merge_stock_df['鐵管建議採購量'] = 採購單_鐵管_merge_stock_df['鐵管建議採購量'].apply(lambda x: x if x > 0 else 0)

# 確保合併後 '最低訂購量' 欄位是數字且非空
採購單_鐵管_merge_stock_df['最低訂購量'] = pd.to_numeric(採購單_鐵管_merge_stock_df['最低訂購量'], errors='coerce').fillna(0)
#採購量計算（比較最低訂購量）
採購單_鐵管_merge_stock_df['鐵管建議採購量'] = 採購單_鐵管_merge_stock_df.apply(
    lambda row: max(row['最低訂購量'], row['鐵管建議採購量']) if row['鐵管建議採購量'] > 0 else 0,
    axis=1
).astype('Int64')

# 鐵管建議採購量欄位資料 COPY　到數量欄位
採購單_鐵管["數量"] = 採購單_鐵管_merge_stock_df['鐵管建議採購量']
# ECOUNT上傳格式 刪除不需要的欄位 
採購單_鐵管.drop(columns=['最低訂購量'], inplace=True, errors='ignore')  # 刪除不需要的欄位





####################################################
#取有設定生管BOM 生產工單
####################################################
生產工單Finalrow = Finalrow[Finalrow['流程種類2'] == '生產工單']
# 只保留光陽訂單需求量不為 0 的資料
生產工單Finalrow= 生產工單Finalrow[生產工單Finalrow['光陽訂單需求量'] > 0]

# 計算每個品項編碼的光陽訂單需求量加總
#生產工單光陽訂單需求量加總 = 生產工單Finalrow.groupby('品項編碼', as_index=False)['光陽訂單需求量'].sum().rename(columns={'光陽訂單需求量': '光陽訂單需求量_合計'})

# 對 Finalrow 以「品項編碼」去重複（保留每個品項一筆，保留第一筆資料）
#生產工單Finalrow = 生產工單Finalrow.drop_duplicates(subset='品項編碼', keep='first')

# 合併光陽訂單需求量彙總回來（以品項編碼對應）
#生產工單Finalrow_加總版 = pd.merge(生產工單Finalrow_去重複,  生產工單光陽訂單需求量加總, on='品項編碼', how='left')

# 將原本的「光陽訂單需求量」欄位改為「光陽訂單需求量_合計」的值
#生產工單Finalrow_加總版['光陽訂單需求量'] = 生產工單Finalrow_加總版['光陽訂單需求量_合計']

# 最後可以選擇是否要刪掉「光陽訂單需求量_合計」欄位
#生產工單Finalrow_加總版.drop(columns=['光陽訂單需求量_合計'], inplace=True)


# 建議採購量計算
#生產工單Finalrow_加總版['建議採購量'] = (
#    生產工單Finalrow_加總版['光陽訂單需求量'] + 生產工單Finalrow_加總版['生產工單未領出'] - 生產工單Finalrow_加總版['庫存數量'] - 生產工單Finalrow_加總版['未進貨(在途量)']
#).fillna(0)
#生產工單Finalrow_加總版['建議採購量'] = 生產工單Finalrow_加總版['建議採購量'].apply(lambda x: x if x > 0 else 0)

#未比較最低採購量   待討論
#生產工單Finalrow_加總版['採購量'] = 生產工單Finalrow_加總版['建議採購量']

#計算交付日期
#生產工單Finalrow['交付日期'] =生產工單Finalrow['供應商工作日'].apply(計算交付日期)

df_生產工單 = pd.DataFrame({
    "日期": today_str,
    "序號": "1",
    "生產數量": 生產工單Finalrow['來源合計'],
    "領料日期": "",
    "建單人員": "",
    "客戶編碼": "21",
    "供應商名": 生產工單Finalrow['製造單位'],
    "工令日期": "",
    "預計領料日": "",
    "母件名稱": 生產工單Finalrow['母件號碼'],
    "番號/規格": "",
    "品項名稱": 生產工單Finalrow['子件號碼'],
    "用量": 生產工單Finalrow['展開後使用量'],
    "數量": 生產工單Finalrow['光陽訂單需求量'],
    "材料批號": "",
    "BOM類型": "",
    "備註1": 生產工單Finalrow['備註1'],
    "備註2": 生產工單Finalrow['備註2']

})
# 依「客戶/供應商名稱」排序
df_生產工單 = df_生產工單.sort_values(by="供應商名").reset_index(drop=True)


# 輸出至 Excel
with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:

    #Finalrow母件路徑.to_excel     (writer, sheet_name='Finalrow母件路徑', index=False)
    Finalrow_原始展開結果.to_excel (writer, sheet_name='Finalrow_原始展開結果', index=False)
    Finalrow_子件號碼排序.to_excel (writer, sheet_name='Finalrow_子件號碼排序', index=False)
    採購單光陽訂單需求量加總.to_excel (writer, sheet_name='採購單光陽訂單需求量加總', index=False)
    採購單Finalrow.to_excel       (writer, sheet_name='採購單Finalrow', index=False)
    採購單Finalrow_加總版Before.to_excel (writer, sheet_name='採購單Finalrow_加總版Before', index=False)
    採購單Finalrow_加總版.to_excel (writer, sheet_name='採購單Finalrow_加總版', index=False)
    df_non_pipe_po.to_excel       (writer, sheet_name='非鐵管_採購單加總版', index=False)


    鐵管Finalrow_original.to_excel(writer, sheet_name='鐵管Finalrow_original', index=False)
    鐵管Finalrow.to_excel         (writer, sheet_name='鐵管Finalrow', index=False)
    鐵管Finalrow_加總版.to_excel   (writer, sheet_name='鐵管Finalrow_加總版', index=False)
    #鐵管Finalrow_2nd加總版.to_excel   (writer, sheet_name='鐵管Finalrow_2nd加總版', index=False)
    採購單_鐵管.to_excel           (writer, sheet_name='採購單_鐵管', index=False)
    採購單_鐵管_merge_stock_df.to_excel      (writer, sheet_name='採購單_鐵管_merge_stock_df', index=False)
    
    生產工單Finalrow.to_excel(writer, sheet_name='生產工單Finalrow', index=False)
    df_生產工單.to_excel           (writer, sheet_name='df_生產工單', index=False)
print("✅ 完成：展開結果已輸出 ")


# 執行時間
timestamp = datetime.now().strftime("執行時間：%Y-%m-%d %H:%M:%S")
timestamp_df = pd.DataFrame([[timestamp]], columns=Finalrow.columns[:1])
Finalrow = pd.concat([
    timestamp_df,
    pd.DataFrame([[''] * len(Finalrow.columns)], columns=Finalrow.columns),
    Finalrow
], ignore_index=True)


# 記錄程式執行結束時間
end_time = time.time()
# 計算執行時間
execution_time = end_time - start_time
log_訊息("1 INFO", f"程式執行時間：{execution_time} 秒")
print(f"程式執行時間：{execution_time} 秒")
input("##### 按任一鍵 跳出程式 #####")
print("......")